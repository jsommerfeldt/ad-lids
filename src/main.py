# main.py
from __future__ import annotations

import logging

from modules.auth import AuthProvider
from modules.config import Config
from modules.graph_client import GraphClient
from modules.query import OneDriveFolderQuery, SQLServerQuery
from modules.summarizer import ThreeWeekSummarizer

import os
from pathlib import Path
import pandas as pd

# Styling imports
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("main")

def generate_output_excel(tws, results):
    # Example: write each folder’s per-file DataFrames to disk (one Excel workbook per folder)
    # Each sheet is a file; DataFrame has a 'SourceFile' column for traceability.

    # Columns to drop from AdLidPriceOnly
    DROP_FROM_AD = ["Product Number", "Type", "Holiday Coloring Code", "Unnamed: 2", "Unnamed: 3"]

    out_dir = Path("assets") / "summaries"
    out_dir.mkdir(parents=True, exist_ok=True)

    for folder_name, files_map in results.items():
        if not files_map:
            continue
        safe_folder = folder_name.replace("/", "_")
        out_path = out_dir / f"{safe_folder}.xlsx"

        with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
            # Manifest (from inventory) for context
            folder_files = tws.resolver.files_under_folder_name(folder_name)
            folder_files.to_excel(xw, index=False, sheet_name="__manifest__")

            # ---- WRITE EACH SHEET ----
            for fname, df_file in files_map.items():
                sheet = fname[:31] or "Sheet"  # Excel sheet name limit

                # If this is the AdLidPriceOnly sheet, drop specified columns if present
                df_to_write = df_file
                if fname == "AdLidPriceOnly":
                    df_to_write = df_file.drop(columns=DROP_FROM_AD, errors="ignore")
                
                # Resort columns
                resort_order = [
                    'Loading Start Date', 'Loading End Date', 'Commodity', 'Vendor', 'Item', 'Description','Ad Lid Price',
                    'FOB or Delivered', 'Confirm by Date', 'Country of Origin', 'Loading Location', 'Estimated Quantity Needed',
                    'Notes', 'Folder', 'SourceFile', 'SheetName']
                df_to_write = df_to_write[resort_order]

                df_to_write.to_excel(xw, index=False, sheet_name=sheet)

            """ OLD
            # ---- AUTOFIT ALL SHEETS (openpyxl) ----
            wb = xw.book
            for ws in wb.worksheets:
                # Determine max width per column using header values
                widths = {}
                # Include header row
                for col_idx, cell in enumerate(ws[1], start=1):
                    txt = str(cell.value) if cell.value is not None else ""
                    widths[col_idx] = max(widths.get(col_idx, 0), len(txt))
                # Include data rows (limit extremely long strings)
                for row in ws.iter_rows(min_row=2):
                    for col_idx, cell in enumerate(row, start=1):
                        val = cell.value
                        if val is None:
                            ln = 0
                        else:
                            # Stringify with a reasonable cap to avoid huge widths
                            s = str(val)
                            if len(s) > 200:
                                s = s[:200]
                            ln = len(s)
                        widths[col_idx] = max(widths.get(col_idx, 0), ln)
                # Apply width with padding and a minimum
                for col_idx, max_len in widths.items():
                    col_letter = get_column_letter(col_idx)
                    # Approximate width: characters + padding; clamp to sensible bounds
                    adjusted = max(8, min(max_len + 2, 60))
                    ws.column_dimensions[col_letter].width = adjusted
            """

            # ---- AUTOFIT ALL SHEETS (openpyxl) ----
            wb = xw.book
            # Columns that should NOT be autofitted
            SKIP_AUTOFIT_COLS = {"Folder", "SourceFile", "SheetName"}

            for ws in wb.worksheets:
                # Build a header map: col_idx -> header_text
                header_map = {}
                for col_idx, cell in enumerate(ws[1], start=1):
                    hdr = "" if cell.value is None else str(cell.value).strip()
                    header_map[col_idx] = hdr

                # Determine max width per column using header values, but skip specified columns
                widths = {}
                # Include header row
                for col_idx, cell in enumerate(ws[1], start=1):
                    hdr = header_map.get(col_idx, "")
                    if hdr in SKIP_AUTOFIT_COLS:
                        continue  # block autofitting for these columns
                    txt = str(cell.value) if cell.value is not None else ""
                    widths[col_idx] = max(widths.get(col_idx, 0), len(txt))

                # Include data rows (limit extremely long strings)
                for row in ws.iter_rows(min_row=2):
                    for col_idx, cell in enumerate(row, start=1):
                        hdr = header_map.get(col_idx, "")
                        if hdr in SKIP_AUTOFIT_COLS:
                            continue  # do not measure widths for these columns
                        val = cell.value
                        if val is None:
                            ln = 0
                        else:
                            s = str(val)
                            if len(s) > 200:
                                s = s[:200]
                            ln = len(s)
                        widths[col_idx] = max(widths.get(col_idx, 0), ln)

                # Apply width with padding and a minimum, skipping blocked columns
                for col_idx, max_len in widths.items():
                    hdr = header_map.get(col_idx, "")
                    if hdr in SKIP_AUTOFIT_COLS:
                        continue  # explicitly do not set width for these columns
                    col_letter = get_column_letter(col_idx)
                    adjusted = max(8, min(max_len + 2, 60))
                    ws.column_dimensions[col_letter].width = adjusted

            # ---- ALTERNATING HIGHLIGHT BY UNIQUE 'Item' GROUPS: ONLY for 'AdLidPriceOnly' ----
            # Toggling fill whenever the 'Item' value changes from the previous row.
            alt_fill_a = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")  # light gray
            alt_fill_b = PatternFill(start_color="FFFFFFFF",  end_color="FFFFFFFF",  fill_type="solid")  # white

            for ws in wb.worksheets:

                if ws.title != "AdLidPriceOnly":
                    continue  # apply styling only to AdLidPriceOnly

                # Find the 'Item' column index by scanning the header row (row 1)
                item_col_idx = None
                for col_idx, cell in enumerate(ws[1], start=1):
                    if str(cell.value).strip().lower() == "item":
                        item_col_idx = col_idx
                        break

                if item_col_idx is None:
                    # No 'Item' column -> nothing to group; leave as-is
                    continue

                # Walk data rows and toggle fill per contiguous Item groups
                current_fill = alt_fill_a
                previous_item = None

                for r in range(2, ws.max_row + 1):
                    item_val = ws.cell(row=r, column=item_col_idx).value

                    if r == 2:
                        # start first data group with alt_fill_a
                        current_fill = alt_fill_a
                    else:
                        if item_val != previous_item:
                            # New group -> toggle fill
                            current_fill = alt_fill_b if current_fill == alt_fill_a else alt_fill_a

                    # Apply fill to the entire row across existing columns
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = current_fill

                    previous_item = item_val

            for ws in wb.worksheets:
                # Hide irrelevant sheets
                if ws.title in ('__manifest__', 'Consolidated'):
                    ws.sheet_state = "hidden"

    return out_dir

def main():
    # 1) Load config and acquire token
    cfg = Config()
    token = AuthProvider(cfg).acquire_token()

    # 2) Build Graph client
    gc = GraphClient(token, base=cfg.GRAPH_BASE)

    # 3) Query helper rooted at the working folder ("Ad Lids")
    q = OneDriveFolderQuery(graph_client=gc, owner_upn=cfg.OWNER_UPN, base_root=cfg.BASE_FOLDER_PATH)

    # 4) Get the inventory table (recursive)
    df = q.to_dataframe(recursive=True, include_files=True, include_folders=True)

    # 5) Sort the data
    df = q.order_by_top_folder_block(df=df, root_position="top")

    # 6) Get dim_time table from SQL Server for determining the relevant directory
    ss = SQLServerQuery()
    engine = ss.get_engine()
    time_df = ss.fetch_data(query=f"""\
    SELECT *
    FROM OPENQUERY(PPRODW,
        '
            SELECT sundayweeknumber
            FROM dim_time dt
            WHERE date = CURRENT_DATE
        '
    )
    """,
    engine=engine)
    sundayweeknumber = int(time_df.loc[0, 'sundayweeknumber'])

    # Save to CSV for inspection
    df.to_csv("assets\\ad_lids_inventory.csv", index=False)

    # Generate output excel file
    tws = ThreeWeekSummarizer(graph_client=gc, owner_upn=cfg.OWNER_UPN, inventory_df=df)
    results = tws.run(sundayweeknumber=sundayweeknumber, horizon=3)
    
    results_consolidated = tws.summarize_books(results)
    out_dir = generate_output_excel(tws=tws, results=results_consolidated)

    logger.info("Summaries written to: %s", out_dir.resolve())

    # Upload each generated .xlsx so that each file lives in a folder named after itself
    cfg = Config()
    gc = GraphClient(AuthProvider(cfg).acquire_token(), base=cfg.GRAPH_BASE)

    for p in Path(out_dir).glob("*.xlsx"):
        gc.upload_local_file_into_same_named_folder(
            upn=cfg.OWNER_UPN,
            local_relative_path=str(p),
            base_folder_path=cfg.BASE_FOLDER_PATH,  # e.g., "Ad Lids"
            folder_name_mode="file-name",           # per your instruction
            conflict_behavior="replace"             # overwrite if the file already exists
        )

if __name__ == "__main__":
    main()
