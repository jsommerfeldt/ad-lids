def generate_output_excel(tws, results):
    # Example: write each folder’s per-file DataFrames to disk (one Excel workbook per folder)
    # Each sheet is a file; DataFrame has a 'SourceFile' column for traceability.
    # Columns to drop from AdLidPriceOnly
    DROP_FROM_AD = ["Product Number", "Type", "Holiday Coloring Code", "Unnamed: 2", "Unnamed: 3"]

    out_dir = Path("assets") / "summaries"
    out_dir.mkdir(parents=True, exist_ok=True)

    for folder_name, files_map in results.items():
        if not files_map:
            continue

        safe_folder = folder_name.replace("/", "_")
        out_path = out_dir / f"{safe_folder}.xlsx"

        with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
            # Manifest (from inventory) for context
            folder_files = tws.resolver.files_under_folder_name(folder_name)
            folder_files.to_excel(xw, index=False, sheet_name="__manifest__")

            # ---- WRITE EACH SHEET ----
            for fname, df_file in files_map.items():
                sheet = fname[:31] or "Sheet"  # Excel sheet name limit

                # If this is the AdLidPriceOnly sheet, drop specified columns if present
                df_to_write = df_file
                if fname == "AdLidPriceOnly":
                    df_to_write = df_file.drop(columns=DROP_FROM_AD, errors="ignore")

                # Resort columns (as you specified)
                resort_order = [
                    'Loading Start Date', 'Loading End Date', 'Commodity', 'Vendor', 'Item', 'Description',
                    'Ad Lid Price', 'FOB or Delivered', 'Confirm by Date', 'Country of Origin',
                    'Loading Location', 'Estimated Quantity Needed', 'Notes', 'Folder', 'SourceFile', 'SheetName'
                ]
                df_to_write = df_to_write[resort_order]

                df_to_write.to_excel(xw, index=False, sheet_name=sheet)

            # === INSERT BUYER COLUMN WITH VLOOKUP (after Description, before Ad Lid Price) ===
            wb = xw.book
            LOOKUP_RANGE = "'https://russdaviswholesale.sharepoint.com/Shared Documents/RussDavisWholesale/Buyers/[Item Master II.xlsx]PD345 (3)'!$A:$K"

            for ws in wb.worksheets:
                # Skip internal sheets
                if ws.title in ("__manifest__",):
                    continue

                # Build header map: name -> column index (1-based)
                header_map = {}
                for col_idx, cell in enumerate(ws[1], start=1):
                    header_text = "" if cell.value is None else str(cell.value).strip()
                    if header_text:
                        header_map[header_text] = col_idx

                # Ensure required headers exist
                if "Description" not in header_map or "Ad Lid Price" not in header_map or "Item" not in header_map:
                    # If any are missing, do not attempt to insert the Buyer column on this sheet
                    continue

                desc_col = header_map["Description"]
                item_col = header_map["Item"]

                # Insert a column immediately after Description
                insert_at = desc_col + 1
                ws.insert_cols(insert_at, amount=1)
                ws.cell(row=1, column=insert_at, value="Buyer")  # header

                # Recompute the item column letter (in case insert shifted columns left/right)
                # We purposely resolve the letter AFTER the insert; item column index doesn't move because it is left of the insert.
                from openpyxl.utils import get_column_letter
                item_col_letter = get_column_letter(item_col)
                buyer_col_letter = get_column_letter(insert_at)

                # Fill formulas for data rows (row 2 .. max_row)
                for r in range(2, ws.max_row + 1):
                    # VLOOKUP on the Item cell in this row
                    formula = f"=VLOOKUP({item_col_letter}{r},{LOOKUP_RANGE},11,0)"
                    ws.cell(row=r, column=insert_at).value = formula
            # === END INSERT BUYER COLUMN ===

            # ---- AUTOFIT ALL SHEETS (openpyxl) ----
            wb = xw.book
            # Columns that should NOT be autofitted
            SKIP_AUTOFIT_COLS = {"Folder", "SourceFile", "SheetName"}
            for ws in wb.worksheets:
                # Build a header map: col_idx -> header_text
                header_map = {}
                for col_idx, cell in enumerate(ws[1], start=1):
                    hdr = "" if cell.value is None else str(cell.value).strip()
                    header_map[col_idx] = hdr

                # Determine max width per column using header values, but skip specified columns
                widths = {}
                # Include header row
                for col_idx, cell in enumerate(ws[1], start=1):
                    hdr = header_map.get(col_idx, "")
                    if hdr in SKIP_AUTOFIT_COLS:
                        continue  # block autofitting for these columns
                    txt = str(cell.value) if cell.value is not None else ""
                    widths[col_idx] = max(widths.get(col_idx, 0), len(txt))
                # Include data rows (limit extremely long strings)
                for row in ws.iter_rows(min_row=2):
                    for col_idx, cell in enumerate(row, start=1):
                        hdr = header_map.get(col_idx, "")
                        if hdr in SKIP_AUTOFIT_COLS:
                            continue  # do not measure widths for these columns
                        val = cell.value
                        if val is None:
                            ln = 0
                        else:
                            s = str(val)
                            if len(s) > 200:
                                s = s[:200]
                            ln = len(s)
                        widths[col_idx] = max(widths.get(col_idx, 0), ln)
                # Apply width with padding and a minimum, skipping blocked columns
                for col_idx, max_len in widths.items():
                    hdr = header_map.get(col_idx, "")
                    if hdr in SKIP_AUTOFIT_COLS:
                        continue  # explicitly do not set width for these columns
                    col_letter = get_column_letter(col_idx)
                    adjusted = max(8, min(max_len + 2, 60))
                    ws.column_dimensions[col_letter].width = adjusted

            # ---- ALTERNATING HIGHLIGHT BY UNIQUE 'Item' GROUPS: ONLY for 'AdLidPriceOnly' ----
            from openpyxl.styles import PatternFill
            alt_fill_a = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")  # light gray
            alt_fill_b = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")  # white
            for ws in wb.worksheets:
                if ws.title != "AdLidPriceOnly":
                    continue  # apply styling only to AdLidPriceOnly
                # Find the 'Item' column index by scanning the header row (row 1)
                item_col_idx = None
                for col_idx, cell in enumerate(ws[1], start=1):
                    if str(cell.value).strip().lower() == "item":
                        item_col_idx = col_idx
                        break
                if item_col_idx is None:
                    # No 'Item' column -> nothing to group; leave as-is
                    continue
                # Walk data rows and toggle fill per contiguous Item groups
                current_fill = alt_fill_a
                previous_item = None
                for r in range(2, ws.max_row + 1):
                    item_val = ws.cell(row=r, column=item_col_idx).value
                    if r == 2:
                        current_fill = alt_fill_a
                    else:
                        if item_val != previous_item:
                            current_fill = alt_fill_b if current_fill == alt_fill_a else alt_fill_a
                    # Apply fill to the entire row across existing columns
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = current_fill
                    previous_item = item_val

            for ws in wb.worksheets:
                # Hide irrelevant sheets
                if ws.title in ('__manifest__', 'Consolidated'):
                    ws.sheet_state = "hidden"

